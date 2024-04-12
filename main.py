import pandas as pd
pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)
import openpyxl
from openpyxl.styles import PatternFill

# Function to assign group to each student
def assign_group(row, group_size):
    group_letter = chr(ord('A') + row.name % group_size)
    return group_letter

if __name__ == '__main__':
    # Read students csv
    students = pd.read_csv('students.csv')
    students = students.rename(columns={"COGNOME - (*) Inserito dal docente": "COGNOME"})
    students['MATRICOLA'] = students['MATRICOLA'].astype(str)

    # Define group dimension
    group_dimension = 7
    course_column = 'CDS STUDENTE'

    num_students = students.shape[0]
    num_groups = num_students // group_dimension
    groups_greater = num_students % num_groups
    print(f'Number of students: {num_students}\nNumber of groups: {num_groups}\nNumber of students in each group: {group_dimension}\nNumber of groups with one more student: {groups_greater}')
    students['course_count'] = students.groupby(course_column)[course_column].transform('count')
    students['course_count_rank'] = students[course_column].rank(method='min', ascending=False)
    print(students[course_column].value_counts())

    # Sort students by course count and assign group
    students = students.sort_values(['course_count_rank',course_column], ascending=True)
    students = students.reset_index(drop=True)
    students['GRUPPO'] = students.apply(assign_group, axis=1, args=(num_groups,))
    print(students[["MATRICOLA",course_column, "course_count", "course_count_rank", "GRUPPO"]])

    # Check diversity
    grouped_students = students.groupby('GRUPPO')[course_column].apply(list)
    print(grouped_students)
    grouped_students_count = students.groupby('GRUPPO')[course_column].nunique()
    print(f"min diversity: {min(grouped_students_count)}")

    #save to csv
    students.to_csv('students-output.csv', index=False)
    #save to xslx
    students[['GRUPPO', 'MATRICOLA', 'COGNOME', 'NOME']].sort_values(['GRUPPO', 'COGNOME'], ascending=True).to_excel('students-output.xlsx', sheet_name="Gruppi P&T 2024", index=False)
    
    # Read the existing Excel file
    workbook = openpyxl.load_workbook('students-output.xlsx')
    sheet = workbook['Gruppi P&T 2024']
    
    # Define the fill colors
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    grey_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

    # Apply the fill colors and borders based on the "GRUPPO" column
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
        group_value = row[0].value
        if ord(group_value) % 2 == 0:
            fill_color = white_fill
        else:
            fill_color = grey_fill
        for cell in row:
            cell.fill = fill_color
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
    # Resize the column width to show all the content
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width
    # Save the modified Excel file
    workbook.save('students-output.xlsx')
